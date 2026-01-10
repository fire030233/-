import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =================================================
# 기본 설정
# =================================================
st.set_page_config(
    page_title="🧪 시약 유통기한 자동 관리",
    layout="wide"
)

FILE_NAME = "reagents.xlsx"

# =================================================
# 데이터 로드
# =================================================
@st.cache_data
def load_data():
    df = pd.read_excel(FILE_NAME)
    df['등록일'] = pd.to_datetime(df['등록일'], errors="coerce")
    df['유통기한'] = pd.to_datetime(df['유통기한'], errors="coerce")
    return df

df = load_data()

# =================================================
# 날짜 계산
# =================================================
today = pd.to_datetime(datetime.today().date())
df['남은일수'] = (df['유통기한'] - today).dt.days
df = df.sort_values(by='남은일수')

# =================================================
# 화면 표시
# =================================================
st.title("🧪 시약 유통기한 자동 관리 시스템")
st.write(f"📅 기준일: **{today.date()}**")

def color_df(row):
    if row['남은일수'] < 0:
        return ['background-color:#ffcccc'] * len(row)
    elif row['남은일수'] <= 30:
        return ['background-color:#fff2cc'] * len(row)
    return ['background-color:white'] * len(row)

# =================================================
# 🔍 시약 제품명 검색
# =================================================
st.subheader("🔍 시약 검색")

search_term = st.text_input("시약 제품명 입력 (부분 검색 가능)")

filtered_df = df.copy()

if search_term:
    filtered_df = filtered_df[
        filtered_df['제품명'].astype(str).str.contains(search_term, case=False, na=False)
    ]

st.dataframe(
    filtered_df.style.apply(color_df, axis=1),
    use_container_width=True
)

# =================================================
# 엑셀 다운로드 (색상 포함)
# =================================================
st.divider()
st.subheader("📥 엑셀 다운로드 (색상 포함)")

if st.button("📥 엑셀 파일 다운로드"):

    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    red = PatternFill("solid", start_color="FFCCCC")
    yellow = PatternFill("solid", start_color="FFF2CC")

    remain_col = [cell.value for cell in ws[1]].index("남은일수") + 1

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=remain_col).value
        if val < 0:
            fill = red
        elif val <= 30:
            fill = yellow
        else:
            continue

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.download_button(
        label="⬇️ 엑셀 파일 저장",
        data=final_output,
        file_name="시약_유통기한_자동관리_결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )